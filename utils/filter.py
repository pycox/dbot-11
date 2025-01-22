from openpyxl import load_workbook
import os
from pymongo import MongoClient
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook

load_dotenv()

db_uri = os.getenv("DB_URI")

client = MongoClient(db_uri)
db = client["dbot"]
collection = db["leads"]

ctrXlDir = r"clients.xlsx"
logDir = rf"logs/{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

rowData = []

os.makedirs(os.path.dirname(logDir), exist_ok=True)


def filterUrls():
    wb = load_workbook(ctrXlDir)

    ws = wb.active

    urls = []

    if ws["D1"].value != "yes":
        return urls

    for row in ws.iter_rows(min_row=2, max_row=601):
        # for row in ws.iter_rows(min_row=1):

        if row[0].value == "ID" or row[0].value is None:
            continue

        if row[3].value == "yes":
            urls.append((row[0].value, row[1].value, row[2].value))

    return urls


def readUrl(key):
    wb = load_workbook(ctrXlDir)

    ws = wb.active

    if isinstance(key, int) and ws[f"A{key + 1}"].value == key:
        return [ws[f"B{key + 1}"].value, ws[f"C{key + 1}"].value]

    for row in ws.iter_rows(min_row=1):
        if str(row[0].value) == str(key):
            return [row[1].value, row[2].value]


def eventHander(key, tp):
    if os.path.exists(logDir):
        wb = load_workbook(logDir)
    else:
        wb = Workbook()

    if wb.active:
        ws = wb.active
    else:
        ws = wb.create_sheet()

    if ws["A1"] != "ID":
        ws["A1"] = "ID"
        ws["B1"] = "Company"
        ws["C1"] = "URL"
        ws["D1"] = "Log"

    company, url = readUrl(key)

    ws.append([key, company, url, tp])

    wb.save(logDir)


def updateDB(key, arr):
    # print(
    #     "########",
    #     f".{key}.",
    #     f"[{len(arr)}]",
    #     arr,
    #     "########",
    #     f".{key}.",
    #     f"[{len(arr)}]",
    # )

    if len(arr) == 0:
        eventHander(key, "EMPTY")
        return

    for item in arr:
        title, company, location, link = item

        result = collection.update_one(
            {"url": link},
            {
                "$set": {
                    "num": key,
                    "url": link,
                    "company": company,
                    "jobTitle": title,
                    "location": location,
                    "createdAt": datetime.now(),
                }
            },
            upsert=True,
        )
