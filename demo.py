from pymongo import MongoClient
from openpyxl import load_workbook, Workbook

ctrXlDir = r"clients.xlsx"
wb = load_workbook(ctrXlDir)
ws = wb.active

client = MongoClient("mongodb://localhost:27017/")
db = client["dbot"]
collection = db["leads"]

leads = collection.find()
nums = {lead["num"] for lead in leads}

filtered_data = []

for row in ws.iter_rows(min_row=2, max_row=601):
    if row[0].value == "ID" or row[0].value is None:
        continue

    if row[3].value == "yes" and row[0].value not in nums:
        filtered_data.append((row[0].value, row[1].value, row[2].value))

new_wb = Workbook()
new_ws = new_wb.active

new_ws.append(("ID", "Client", "Url"))

for data in filtered_data:
    new_ws.append(data)

new_wb.save("demo.xlsx")
